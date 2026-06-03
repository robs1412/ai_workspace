#!/usr/bin/env python3

from __future__ import annotations

import csv
import datetime as dt
import html
import io
import os
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from PIL import Image
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import black, white
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


OLD_K_RE = re.compile(r"^(K_\d{3})-(.+)$")
REF_K_RE = re.compile(r"\bK_(\d{3})\b")
SMALL_REMOVE_DIMS = {(35, 35), (43, 43), (128, 38), (129, 38)}


@dataclass
class SourceItem:
    old_k: str
    old_name: str
    old_path: Path
    suffix_name: str
    ext: str
    row: dict[str, str]
    remove_reason: str | None = None


@dataclass
class OutputItem:
    old_k: str
    old_name: str
    new_name: str
    start_k: str
    end_k: str
    start_page: int
    end_page: int
    pages: int
    month: str
    doc_type: str
    title: str
    description: str
    from_field: str
    to_field: str


def run(cmd: list[str], *, cwd: Path | None = None, check: bool = True) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        cmd,
        cwd=str(cwd) if cwd else None,
        check=check,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )


def k_label(page_num: int) -> str:
    return f"K{page_num:04d}"


def normalize_spaces(text: str) -> str:
    return " ".join((text or "").split())


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def build_source_items(source_dir: Path, index_rows: list[dict[str, str]]) -> list[SourceItem]:
    row_by_file = {row["File"]: row for row in index_rows if row.get("File")}
    items: list[SourceItem] = []
    for path in sorted(source_dir.iterdir()):
        if not path.is_file():
            continue
        if path.name.startswith("."):
            continue
        match = OLD_K_RE.match(path.name)
        if not match:
            continue
        old_k, suffix_name = match.groups()
        row = row_by_file.get(path.name, {})
        items.append(
            SourceItem(
                old_k=old_k,
                old_name=path.name,
                old_path=path,
                suffix_name=suffix_name,
                ext=path.suffix.lower(),
                row=row,
            )
        )
    return items


def classify_removals(items: Iterable[SourceItem]) -> None:
    for item in items:
        if item.ext not in {".png", ".jpg", ".jpeg"}:
            continue
        try:
            with Image.open(item.old_path) as img:
                dims = img.size
        except Exception:
            continue
        if dims in SMALL_REMOVE_DIMS or (dims[0] <= 45 and dims[1] <= 45):
            item.remove_reason = f"small logo/icon attachment ({dims[0]}x{dims[1]})"


def reorder_items(items: list[SourceItem]) -> list[SourceItem]:
    by_k = {item.old_k: item for item in items}
    ordered = items[:]

    if "K_001" in by_k and "K_165" in by_k:
        lone_941 = by_k["K_001"]
        ordered = [item for item in ordered if item.old_k != "K_001"]
        insert_at = next(
            (idx for idx, item in enumerate(ordered) if item.old_k == "K_165"),
            len(ordered),
        )
        ordered.insert(insert_at, lone_941)
    return ordered


def write_text_pdf(txt_path: Path, out_pdf: Path) -> None:
    cmd = ["/usr/sbin/cupsfilter", "-m", "application/pdf", str(txt_path)]
    proc = subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out_pdf.write_bytes(proc.stdout)


def write_image_pdf(image_path: Path, out_pdf: Path) -> None:
    page_w, page_h = letter
    margin = 36
    with Image.open(image_path) as img:
        img.load()
        img_w, img_h = img.size
        img_reader = ImageReader(img)

        scale = min((page_w - 2 * margin) / img_w, (page_h - 2 * margin) / img_h)
        draw_w = img_w * scale
        draw_h = img_h * scale
        x = (page_w - draw_w) / 2
        y = (page_h - draw_h) / 2

        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=letter)
        c.drawImage(img_reader, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
        c.showPage()
        c.save()
        out_pdf.write_bytes(packet.getvalue())


def chrome_print(html_path: Path, out_pdf: Path, orientation: str = "portrait") -> None:
    renderer = Path("/Users/werkstatt/ai_workspace/.pdf_renderer/html_to_pdf.js")
    cmd = [
        "node",
        str(renderer),
        str(html_path),
        str(out_pdf),
        orientation,
    ]
    run(cmd)


def convert_docx(docx_path: Path, out_pdf: Path, temp_dir: Path) -> None:
    html_path = temp_dir / f"{docx_path.stem}.html"
    run(["/usr/bin/textutil", "-convert", "html", "-output", str(html_path), str(docx_path)])
    chrome_print(html_path, out_pdf, "portrait")


def clean_number_format(fmt: str) -> str:
    fmt = fmt or "General"
    fmt = re.sub(r'".*?"', '', fmt)
    fmt = re.sub(r"\[[^\]]+\]", "", fmt)
    fmt = fmt.replace("\\", "")
    fmt = fmt.replace("_", "")
    fmt = fmt.replace("*", "")
    return fmt.strip()


def format_general_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(round(value)))
    return f"{value:.15g}"


def excel_display_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, (dt.datetime, dt.date)):
        fmt = clean_number_format(cell.number_format).lower()
        if "d-mmm" in fmt:
            return value.strftime("%-d-%b")
        if "m/d/yy" in fmt or "m/d/yyyy" in fmt:
            return value.strftime("%-m/%-d/%y")
        return value.strftime("%Y-%m-%d")

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, (int, float)):
        fmt = clean_number_format(cell.number_format)
        if fmt.lower() == "general":
            return format_general_number(float(value))

        negative = value < 0
        abs_value = abs(float(value))
        has_comma = "," in fmt
        decimal_match = re.search(r"\.([0#]+)", fmt)
        decimals = len(decimal_match.group(1)) if decimal_match else 0
        currency = "$" in cell.number_format
        percent = "%" in fmt

        if percent:
            abs_value *= 100

        number_text = f"{abs_value:,.{decimals}f}" if has_comma else f"{abs_value:.{decimals}f}"
        if decimals == 0 and number_text.endswith(".0"):
            number_text = number_text[:-2]

        if currency:
            number_text = f"${number_text}"
        if percent:
            number_text = f"{number_text}%"
        if negative:
            number_text = f"({number_text})" if "(" in cell.number_format else f"-{number_text}"
        return number_text

    return str(value)


def html_cell(cell) -> str:
    return html.escape(excel_display_value(cell))


def worksheet_bounds(ws) -> tuple[int, int, int, int]:
    min_r = min_c = 10**9
    max_r = max_c = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None or str(value).strip() == "":
                continue
            min_r = min(min_r, cell.row)
            min_c = min(min_c, cell.column)
            max_r = max(max_r, cell.row)
            max_c = max(max_c, cell.column)
    if max_r == 0:
        return 1, 1, 1, 1
    return min_r, min_c, max_r, max_c


def xlsx_to_html(xlsx_path: Path, html_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    parts = [
        "<!doctype html>",
        '<html lang="en"><head><meta charset="utf-8">',
        f"<title>{html.escape(xlsx_path.name)}</title>",
        "<style>",
        "@page { size: landscape; margin: 0.35in; }",
        "body{font-family:Arial,sans-serif;color:#111;margin:0}",
        "h2{font-size:13px;margin:0 0 8px}",
        ".sheet{page-break-after:always;margin-bottom:12px}",
        ".chunk{page-break-after:always;margin-bottom:12px}",
        "table{border-collapse:collapse;width:100%;table-layout:fixed}",
        "th,td{border:1px solid #999;padding:2px 4px;font-size:7px;vertical-align:top;word-break:break-word}",
        "thead th{background:#f1f1f1;font-weight:700}",
        ".titlecell{background:#e7eef8;font-weight:700}",
        ".repeat{background:#fafafa}",
        "</style></head><body>",
    ]

    for ws in wb.worksheets:
        min_r, min_c, max_r, max_c = worksheet_bounds(ws)
        visible_rows = []
        for r in range(min_r, max_r + 1):
            if ws.row_dimensions[r].hidden:
                continue
            vals = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
            if any(v not in (None, "") for v in vals):
                visible_rows.append(r)
        visible_cols = []
        data_rows_for_cols = visible_rows[2:] if len(visible_rows) > 2 else visible_rows
        for c in range(min_c, max_c + 1):
            letter = ws.cell(1, c).column_letter
            if ws.column_dimensions[letter].hidden:
                continue
            vals = [ws.cell(r, c).value for r in data_rows_for_cols]
            if any(v not in (None, "") for v in vals):
                visible_cols.append(c)
        if not visible_rows or not visible_cols:
            continue

        all_cols = visible_cols
        repeat_cols = all_cols[:1]
        remaining_cols = all_cols[1:]
        chunk_size = 17 if len(all_cols) <= 18 else 9
        chunk_sets: list[list[int]] = []
        if remaining_cols:
            for start in range(0, len(remaining_cols), chunk_size):
                chunk_sets.append(repeat_cols + remaining_cols[start:start + chunk_size])
        else:
            chunk_sets.append(repeat_cols)

        for idx, colset in enumerate(chunk_sets, start=1):
            parts.append('<section class="chunk">')
            title = ws.title if len(chunk_sets) == 1 else f"{ws.title} ({idx}/{len(chunk_sets)})"
            parts.append(f"<h2>{html.escape(title)}</h2>")
            parts.append("<table><thead>")
            header_rows = visible_rows[:2]
            for r in header_rows:
                parts.append("<tr>")
                for c in colset:
                    cls = "titlecell" if r == header_rows[0] else ""
                    parts.append(f'<th class="{cls}">{html_cell(ws.cell(r, c))}</th>')
                parts.append("</tr>")
            parts.append("</thead><tbody>")

            data_rows = visible_rows[len(header_rows):]
            for r in data_rows:
                parts.append("<tr>")
                for c in colset:
                    cls = "repeat" if c in repeat_cols else ""
                    parts.append(f'<td class="{cls}">{html_cell(ws.cell(r, c))}</td>')
                parts.append("</tr>")
            parts.append("</tbody></table></section>")

    parts.append("</body></html>")
    html_path.write_text("\n".join(parts), encoding="utf-8")


def convert_xlsx(xlsx_path: Path, out_pdf: Path, temp_dir: Path) -> None:
    html_path = temp_dir / f"{xlsx_path.stem}.html"
    xlsx_to_html(xlsx_path, html_path)
    chrome_print(html_path, out_pdf, "landscape")


def ensure_pdf_for_item(item: SourceItem, temp_dir: Path) -> Path:
    out_pdf = temp_dir / f"{item.old_k}.pdf"
    if item.ext == ".pdf":
        shutil.copy2(item.old_path, out_pdf)
    elif item.ext in {".png", ".jpg", ".jpeg"}:
        write_image_pdf(item.old_path, out_pdf)
    elif item.ext == ".txt":
        write_text_pdf(item.old_path, out_pdf)
    elif item.ext == ".docx":
        convert_docx(item.old_path, out_pdf, temp_dir)
    elif item.ext == ".xlsx":
        convert_xlsx(item.old_path, out_pdf, temp_dir)
    else:
        raise ValueError(f"Unsupported source type: {item.old_name}")
    return out_pdf


def build_stamp_overlay(width: float, height: float, label: str) -> PdfReader:
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(width, height))
    c.setFillColor(white)
    c.rect((width / 2) - 34, 10, 68, 14, stroke=0, fill=1)
    c.setFillColor(black)
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, 14, label)
    c.showPage()
    c.save()
    packet.seek(0)
    return PdfReader(packet)


def stamp_pdf(input_pdf: Path, out_pdf: Path, starting_page_num: int) -> int:
    reader = PdfReader(str(input_pdf))
    writer = PdfWriter()
    page_num = starting_page_num
    for page in reader.pages:
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        overlay_page = build_stamp_overlay(width, height, k_label(page_num)).pages[0]
        page.merge_page(overlay_page)
        writer.add_page(page)
        page_num += 1
    with out_pdf.open("wb") as handle:
        writer.write(handle)
    return len(reader.pages)


def split_suffix_name(item: SourceItem) -> tuple[str, str]:
    stem, ext = os.path.splitext(item.suffix_name)
    return stem, ext


def new_output_name(item: SourceItem, start_page: int, end_page: int) -> str:
    stem, ext = split_suffix_name(item)
    prefix = f"{k_label(start_page)}-{k_label(end_page)}"
    if ext.lower() == ".pdf":
        return f"{prefix}-{stem}.pdf"
    return f"{prefix}-{stem}{ext}.pdf"


def render_index_html(rows: list[OutputItem], html_path: Path, folder_path: Path) -> None:
    lines = [
        "<!doctype html>",
        '<html lang="en"><head><meta charset="utf-8">',
        "<title>Discovery Production One-Folder New Index</title>",
        "<style>",
        "body{font-family:-apple-system,BlinkMacSystemFont,\"Segoe UI\",Arial,sans-serif;margin:24px;color:#111}",
        "h1{font-size:22px;margin:0 0 8px}",
        ".meta{font-size:13px;color:#555;margin:0 0 18px;line-height:1.4}",
        "table{border-collapse:collapse;width:100%;font-size:11px;line-height:1.3}",
        "th{background:#f3f5f7;border:1px solid #d8dde3;padding:6px;text-align:left}",
        "td{border:1px solid #e1e4e8;padding:6px;vertical-align:top}",
        "tr:nth-child(even){background:#fbfbfc}",
        ".num{white-space:nowrap}",
        ".file{font-family:Menlo,Consolas,monospace;font-size:10px}",
        "a{color:#0645ad;text-decoration:none}a:hover{text-decoration:underline}",
        "</style></head><body>",
        "<h1>Discovery Production One-Folder New Index</h1>",
        (
            f"<p class=\"meta\">Folder: <code>{html.escape(str(folder_path))}</code><br>"
            f"Rows: {len(rows)}. Every retained page is Bates-numbered sequentially.</p>"
        ),
        "<table><thead><tr><th>Start</th><th>End</th><th>Month</th><th>Type</th><th>File</th><th>Subject / Title</th><th>Original K</th></tr></thead><tbody>",
    ]
    for row in rows:
        href = "file://" + str(folder_path / row.new_name)
        lines.append(
            "<tr>"
            f"<td class=\"num\">{row.start_k}</td>"
            f"<td class=\"num\">{row.end_k}</td>"
            f"<td>{html.escape(row.month)}</td>"
            f"<td>{html.escape(row.doc_type)}</td>"
            f"<td class=\"file\"><a href=\"{html.escape(href)}\">{html.escape(row.new_name)}</a></td>"
            f"<td>{html.escape(row.title)}</td>"
            f"<td class=\"num\">{html.escape(row.old_k)}</td>"
            "</tr>"
        )
    lines.append("</tbody></table></body></html>")
    html_path.write_text("\n".join(lines), encoding="utf-8")


def render_suggestion_html(rows: list[dict[str, str]], html_path: Path) -> None:
    lines = [
        "<!doctype html>",
        '<html lang="en"><head><meta charset="utf-8">',
        "<title>Discovery Response Request Crosswalk (Renumbered)</title>",
        "<style>",
        "body{font-family:-apple-system,BlinkMacSystemFont,\"Segoe UI\",Arial,sans-serif;margin:24px;color:#111}",
        "h1{font-size:22px;margin:0 0 8px}",
        ".meta{font-size:13px;color:#555;margin:0 0 18px;line-height:1.4}",
        "table{border-collapse:collapse;width:100%;font-size:11px;line-height:1.3}",
        "th{background:#f3f5f7;border:1px solid #d8dde3;padding:6px;text-align:left}",
        "td{border:1px solid #e1e4e8;padding:6px;vertical-align:top}",
        "tr:nth-child(even){background:#fbfbfc}",
        ".num{white-space:nowrap;text-align:right}",
        "</style></head><body>",
        "<h1>Discovery Response Request Crosswalk (Renumbered)</h1>",
        "<p class=\"meta\">Updated from the v1 request crosswalk, replacing old document-level K identifiers with the new page-level Bates ranges.</p>",
        "<table><thead><tr><th>No.</th><th>Request</th><th>Draft Response</th><th>Documents Produced / K Numbers</th><th>Mapping Note</th></tr></thead><tbody>",
    ]
    for row in rows:
        lines.append(
            "<tr>"
            f"<td class=\"num\">{html.escape(row['Request No.'])}</td>"
            f"<td>{html.escape(row['Request'])}</td>"
            f"<td>{html.escape(row['Draft Response'])}</td>"
            f"<td>{html.escape(row['Documents Produced / K Numbers'])}</td>"
            f"<td>{html.escape(row['Mapping Note'])}</td>"
            "</tr>"
        )
    lines.append("</tbody></table></body></html>")
    html_path.write_text("\n".join(lines), encoding="utf-8")


def update_crosswalk_rows(
    rows: list[dict[str, str]],
    output_map: dict[str, OutputItem],
) -> list[dict[str, str]]:
    updated: list[dict[str, str]] = []
    for row in rows:
        raw = row.get("Documents Produced / K Numbers", "")
        if raw.strip().lower() == "none listed":
            updated.append(row.copy())
            continue

        seen: set[str] = set()
        replacements: list[str] = []
        missing: list[str] = []
        for match in REF_K_RE.finditer(raw):
            old_k = f"K_{match.group(1)}"
            mapped = output_map.get(old_k)
            if not mapped:
                if old_k not in missing:
                    missing.append(old_k)
                continue
            label = mapped.start_k if mapped.start_k == mapped.end_k else f"{mapped.start_k}-{mapped.end_k}"
            if label not in seen:
                seen.add(label)
                replacements.append(label)

        new_row = row.copy()
        new_row["Documents Produced / K Numbers"] = ", ".join(replacements) if replacements else "None listed"
        if missing:
            note = new_row.get("Mapping Note", "").rstrip()
            suffix = f" Omitted missing/removed refs: {', '.join(missing)}."
            if suffix not in note:
                new_row["Mapping Note"] = (note + suffix).strip()
        updated.append(new_row)
    return updated


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def output_index_csv_rows(rows: list[OutputItem]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in rows:
        out.append(
            {
                "Start K": row.start_k,
                "End K": row.end_k,
                "Month": row.month,
                "Type": row.doc_type,
                "File": row.new_name,
                "Subject / Title": row.title,
                "Description": row.description,
                "From": row.from_field,
                "To": row.to_field,
                "Original K": row.old_k,
                "Original File": row.old_name,
            }
        )
    return out


def archive_inputs(
    source_items: list[SourceItem],
    retained_old_ks: set[str],
    source_dir: Path,
) -> tuple[Path, Path, list[dict[str, str]]]:
    removed_dir = source_dir / "__removed_from_production"
    preserved_dir = source_dir / "__original_nonpdf_sources"
    removed_dir.mkdir(exist_ok=True)
    preserved_dir.mkdir(exist_ok=True)

    removed_manifest: list[dict[str, str]] = []
    for item in source_items:
        if item.remove_reason:
            destination = removed_dir / item.old_name
            if destination.exists():
                destination.unlink()
            shutil.move(str(item.old_path), str(destination))
            removed_manifest.append(
                {
                    "Old K": item.old_k,
                    "Old File": item.old_name,
                    "Reason": item.remove_reason,
                }
            )
        elif item.old_k in retained_old_ks and item.ext != ".pdf":
            destination = preserved_dir / item.old_name
            if destination.exists():
                destination.unlink()
            shutil.move(str(item.old_path), str(destination))

    return removed_dir, preserved_dir, removed_manifest


def clear_existing_outputs(source_dir: Path) -> None:
    for path in source_dir.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("K") and not path.name.startswith("K_"):
            path.unlink()


def main(argv: list[str]) -> int:
    if len(argv) != 4:
        print(
            "Usage: ertc_renumber.py <source_dir> <v1_index_csv> <v1_suggestion_csv>",
            file=sys.stderr,
        )
        return 2

    source_dir = Path(argv[1]).resolve()
    v1_index_csv = Path(argv[2]).resolve()
    v1_suggestion_csv = Path(argv[3]).resolve()
    parent_dir = source_dir.parent

    index_rows = load_csv_rows(v1_index_csv)
    suggestion_rows = load_csv_rows(v1_suggestion_csv)
    source_items = build_source_items(source_dir, index_rows)
    classify_removals(source_items)

    retained_items = [item for item in source_items if not item.remove_reason]
    retained_items = reorder_items(retained_items)

    with tempfile.TemporaryDirectory(prefix="ertc-renumber-") as tmp:
        tmp_dir = Path(tmp)
        stage_dir = tmp_dir / "staged"
        stage_dir.mkdir()

        output_rows: list[OutputItem] = []
        page_counter = 1
        for item in retained_items:
            normalized_pdf = ensure_pdf_for_item(item, tmp_dir)
            reader = PdfReader(str(normalized_pdf))
            pages = len(reader.pages)
            start_page = page_counter
            end_page = page_counter + pages - 1
            new_name = new_output_name(item, start_page, end_page)
            stamped_pdf = stage_dir / new_name
            stamp_pdf(normalized_pdf, stamped_pdf, start_page)

            row = item.row
            output_rows.append(
                OutputItem(
                    old_k=item.old_k,
                    old_name=item.old_name,
                    new_name=new_name,
                    start_k=k_label(start_page),
                    end_k=k_label(end_page),
                    start_page=start_page,
                    end_page=end_page,
                    pages=pages,
                    month=row.get("Month", ""),
                    doc_type=row.get("Type", ""),
                    title=row.get("Subject / Title", "") or item.suffix_name,
                    description=row.get("Description", ""),
                    from_field=row.get("From", ""),
                    to_field=row.get("To", ""),
                )
            )
            page_counter = end_page + 1

        output_map = {row.old_k: row for row in output_rows}
        updated_crosswalk = update_crosswalk_rows(suggestion_rows, output_map)

        clear_existing_outputs(source_dir)
        retained_old_ks = {row.old_k for row in output_rows}
        removed_dir, preserved_dir, removed_manifest = archive_inputs(source_items, retained_old_ks, source_dir)

        for stamped_pdf in sorted(stage_dir.iterdir()):
            destination = source_dir / stamped_pdf.name
            if destination.exists():
                destination.unlink()
            shutil.move(str(stamped_pdf), str(destination))

    index_csv_rows = output_index_csv_rows(output_rows)
    index_csv = parent_dir / "discovery-production-one-folder-new-index.csv"
    index_html = parent_dir / "discovery-production-one-folder-new-index.html"
    suggestion_csv = parent_dir / "discovery-response-request-suggestion-renumbered.csv"
    suggestion_html = parent_dir / "discovery-response-request-suggestion-renumbered.html"
    removed_manifest_csv = source_dir / "__removed_from_production_manifest.csv"

    write_csv(
        index_csv,
        index_csv_rows,
        [
            "Start K",
            "End K",
            "Month",
            "Type",
            "File",
            "Subject / Title",
            "Description",
            "From",
            "To",
            "Original K",
            "Original File",
        ],
    )
    render_index_html(output_rows, index_html, source_dir)

    write_csv(
        suggestion_csv,
        updated_crosswalk,
        [
            "Request No.",
            "Request",
            "Draft Response",
            "Documents Produced / K Numbers",
            "Mapping Note",
        ],
    )
    render_suggestion_html(updated_crosswalk, suggestion_html)

    write_csv(
        removed_manifest_csv,
        removed_manifest,
        ["Old K", "Old File", "Reason"],
    )

    print(f"Retained files: {len(output_rows)}")
    print(f"Removed files: {len(removed_manifest)}")
    print(f"Total pages: {output_rows[-1].end_page if output_rows else 0}")
    print(f"Source originals preserved in: {preserved_dir}")
    print(f"Removed items moved to: {removed_dir}")
    print(f"Index CSV: {index_csv}")
    print(f"Index HTML: {index_html}")
    print(f"Crosswalk CSV: {suggestion_csv}")
    print(f"Crosswalk HTML: {suggestion_html}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
